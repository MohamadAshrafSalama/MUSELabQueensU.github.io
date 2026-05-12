#!/usr/bin/env python3
"""
update_site.py — apply edits from the updates/ folder to the live site.

Workflow:
    1. Edit Excel sheets in updates/
    2. Run: python scripts/update_site.py
    3. Commit the changes git shows you.

On first run, the script auto-creates blank templates in updates/.
On subsequent runs, it reads the Excel sheets and rebuilds the data files
and the auto-generated sections of members.html and index.html.
"""

import argparse
import json
import re
import shutil
import sys
import unicodedata
from datetime import date
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.comments import Comment
except ImportError:
    sys.exit(
        "openpyxl is not installed.\n"
        "Run: python3 -m venv .venv && .venv/bin/pip install openpyxl"
    )

ROOT = Path(__file__).resolve().parent.parent
UPDATES = ROOT / "updates"
DATA = ROOT / "data"
PHOTOS = ROOT / "photos"
FILES = ROOT / "files"

PUBS_DIR = UPDATES / "publications"
PUBS_XLSX = PUBS_DIR / "publications.xlsx"
PUBS_FILES_DIR = PUBS_DIR / "files"

NEWS_DIR = UPDATES / "news"
NEWS_XLSX = NEWS_DIR / "news.xlsx"
NEWS_IMAGES_DIR = NEWS_DIR / "images"

MEMBERS_DIR = UPDATES / "members"
MEMBER_TEMPLATE_DIR = MEMBERS_DIR / "_template"

PUBS_JSON = DATA / "publications.json"
NEWS_JSON = DATA / "news.json"
MEMBERS_JSON = DATA / "members.json"

INDEX_HTML = ROOT / "index.html"
MEMBERS_HTML = ROOT / "members.html"

TOPIC_CODES = ["oss", "edi", "msr", "productivity", "ml4se", "tools", "ese"]
TOPIC_LABELS = {
    "oss": "Open Source",
    "edi": "EDI",
    "msr": "Mining Software Repos",
    "productivity": "Developer Productivity",
    "ml4se": "ML for SE",
    "tools": "Tools",
    "ese": "Empirical SE",
}
TYPE_CODES = ["journal", "conference", "book-chapter", "preprint"]
STATUS_CODES = ["PI", "Graduate", "Undergraduate", "Alumni"]

HEADER_FILL = PatternFill(start_color="102040", end_color="102040", fill_type="solid")
REQUIRED_FILL = PatternFill(start_color="FFE9CC", end_color="FFE9CC", fill_type="solid")
OPTIONAL_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=12)
INSTR_FONT = Font(italic=True, color="555555", size=10)
THIN = Side(border_style="thin", color="C0C0C0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# --------------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------------

def slugify(text):
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^a-zA-Z0-9]+", "-", text).strip("-").lower()
    return text or "item"


def split_list(value):
    if value is None:
        return []
    if isinstance(value, list):
        return [str(v).strip() for v in value if str(v).strip()]
    return [v.strip() for v in str(value).split(",") if v.strip()]


def load_json(path, default):
    if not path.exists():
        return default
    with open(path) as f:
        return json.load(f)


def write_json(path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
        f.write("\n")


def style_header_row(ws, headers, required_indexes):
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        if (col_idx - 1) in required_indexes:
            cell.value = title + " *"
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"


def autosize(ws, widths):
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def replace_between_markers(text, start_marker, end_marker, new_block):
    pattern = re.compile(
        re.escape(start_marker) + r".*?" + re.escape(end_marker),
        re.DOTALL,
    )
    if not pattern.search(text):
        raise SystemExit(
            f"Could not find markers {start_marker!r} ... {end_marker!r} in the HTML. "
            "Add them once and the script will keep replacing the content between them."
        )
    return pattern.sub(start_marker + "\n" + new_block + "\n" + end_marker, text)


# --------------------------------------------------------------------------
# Template creation
# --------------------------------------------------------------------------

def init_publications_template():
    if PUBS_XLSX.exists():
        return
    PUBS_DIR.mkdir(parents=True, exist_ok=True)
    PUBS_FILES_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Publications"

    headers = [
        "Title", "Authors", "Year", "Venue", "Venue Acronym",
        "Type", "Topics", "DOI",
        "PDF URL", "Slides URL", "Abstract",
    ]
    required = {0, 1, 2, 3, 5, 6}
    style_header_row(ws, headers, required)
    autosize(ws, [55, 35, 8, 45, 14, 14, 30, 30, 35, 35, 60])

    # Notes on header cells
    ws["A1"].comment = Comment("Full paper title. Required.", "Site")
    ws["B1"].comment = Comment(
        "Comma-separated, in citation order. Use full names. Example:\n"
        "Mohamed Ouf, Mariam Guizani, Amr Mohamed",
        "Site",
    )
    ws["C1"].comment = Comment("Four-digit year. Example: 2026", "Site")
    ws["D1"].comment = Comment(
        "Full venue name. Example: IEEE/ACM 48th International Conference on Software Engineering",
        "Site",
    )
    ws["E1"].comment = Comment("Short acronym. Example: ICSE, TOSEM, CSCW", "Site")
    ws["F1"].comment = Comment(
        "Pick one from the dropdown:\n" + ", ".join(TYPE_CODES), "Site"
    )
    ws["G1"].comment = Comment(
        "Comma-separated. Pick from:\n" + ", ".join(
            f"{c} ({TOPIC_LABELS[c]})" for c in TOPIC_CODES
        ) + "\nExample: oss, msr",
        "Site",
    )
    ws["H1"].comment = Comment(
        "DOI without https://doi.org/ prefix. Example: 10.1145/3744916.3787782", "Site"
    )
    ws["I1"].comment = Comment(
        "Direct link to preprint PDF (arXiv, personal site, etc.). Optional.", "Site"
    )
    ws["J1"].comment = Comment(
        "Link to slides. If you have a local PDF, drop it in updates/publications/files/ "
        "and put just the filename here (e.g. icse-2026-slides.pdf).", "Site"
    )
    ws["K1"].comment = Comment("Paper abstract. Optional but recommended.", "Site")

    # Data validation: type dropdown
    type_dv = DataValidation(
        type="list", formula1='"' + ",".join(TYPE_CODES) + '"', allow_blank=True
    )
    type_dv.error = "Pick one of: " + ", ".join(TYPE_CODES)
    type_dv.errorTitle = "Invalid type"
    ws.add_data_validation(type_dv)
    type_dv.add(f"F2:F1000")

    # Year validation
    year_dv = DataValidation(type="whole", operator="between", formula1=1990, formula2=2100)
    year_dv.error = "Year must be between 1990 and 2100"
    ws.add_data_validation(year_dv)
    year_dv.add("C2:C1000")

    # Instructions sheet
    instr = wb.create_sheet("Instructions")
    instr["A1"] = "How to add a publication"
    instr["A1"].font = Font(bold=True, size=14)
    rows = [
        "",
        "1. Open the Publications sheet.",
        "2. Add one row per paper. Columns marked with * are required.",
        "3. For Type, click the cell to use the dropdown.",
        "4. For Topics, enter a comma-separated list using these codes:",
        "      " + ", ".join(f"{c} = {TOPIC_LABELS[c]}" for c in TOPIC_CODES),
        "5. For Slides, drop the PDF in updates/publications/files/ and enter just the filename.",
        "6. Save the file.",
        "7. From the project root, run:",
        "      python scripts/update_site.py",
        "8. Review what changed with `git status` and `git diff`, then commit.",
        "",
        "Duplicate detection: rows are merged with existing publications by DOI or by title.",
        "Editing an existing paper: change the row whose DOI matches and run the script.",
    ]
    for i, r in enumerate(rows, start=2):
        instr.cell(row=i, column=1, value=r).font = INSTR_FONT
    instr.column_dimensions["A"].width = 100

    wb.save(PUBS_XLSX)
    print(f"  created {PUBS_XLSX.relative_to(ROOT)}")


def init_news_template():
    if NEWS_XLSX.exists():
        return
    NEWS_DIR.mkdir(parents=True, exist_ok=True)
    NEWS_IMAGES_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "News"
    headers = ["Date Badge", "Sort Date", "Text (HTML allowed)", "Link URL", "Link Label", "Image Filename", "Image Alt"]
    required = {0, 1, 2}
    style_header_row(ws, headers, required)
    autosize(ws, [14, 14, 90, 40, 24, 24, 35])

    ws["A1"].comment = Comment(
        "Short label that appears in the badge. Example: May 2026, Apr 2026, Sep 2025", "Site"
    )
    ws["B1"].comment = Comment(
        "Used for sorting only. Format: YYYY-MM. Newest first will appear at the top.", "Site"
    )
    ws["C1"].comment = Comment(
        "The news body. You can include basic HTML tags like <strong> and <em>.", "Site"
    )
    ws["D1"].comment = Comment("Optional link URL (preprint, doi, etc.).", "Site")
    ws["E1"].comment = Comment(
        'Text shown on the link. Example: "Read the preprint", "Read the paper".', "Site"
    )
    ws["F1"].comment = Comment(
        "Optional. Drop your image in updates/news/images/ and enter just the filename here.",
        "Site",
    )
    ws["G1"].comment = Comment("Required if image is set. Alt text for accessibility.", "Site")

    # Pre-populate with the existing news so the user sees what to do
    existing = [
        ("May 2026", "2026-05", 'Our paper <em>"Do Good, Stay Longer?"</em> by Mohamed Ouf, Amr Mohamed, and Mariam Guizani has been accepted at <strong>EASE 2026</strong> (Research Track). We studied how newcomers become core contributors in open source projects focused on social good.', "https://arxiv.org/abs/2601.23142", "Read the preprint", "", ""),
        ("May 2026", "2026-05", 'Our paper <em>"Same Project, Different Start"</em> by Mohamed Ouf and Mariam Guizani has also been accepted at <strong>EASE 2026</strong> (Research Track). This work looks at how contribution events like hackathons and mentorships shape long-term activity in open source.', "https://arxiv.org/abs/2604.22120", "Read the preprint", "", ""),
        ("Apr 2026", "2026-04", "The MUSE Lab attended <strong>ICSE 2026</strong> in Rio de Janeiro! We presented three papers across the Main Track, NIER, and CHASE.", "", "", "icse-2026-lab.jpg", "Yawar Ashraf, Dr. Mariam Guizani, and Aaliyah Chang at ICSE 2026 in Rio de Janeiro"),
        ("Apr 2026", "2026-04", 'Dr. Guizani and Nathan Cassee organized a fishbowl-style panel at <strong>CHASE 2026</strong> called <em>"The Human Side of Developer Experience"</em>, bringing together researchers for an open discussion on human factors in software engineering.', "", "", "", ""),
        ("Mar 2026", "2026-03", "Three papers accepted at <strong>ICSE 2026</strong>! Our work on OSS4SG community patterns got into the Main Track, our collaboration on LLM-driven documentation translation into the NIER Track, and Aaliyah Chang's work on motivations in software engineering into CHASE.", "", "", "", ""),
        ("Mar 2026", "2026-03", "Our <em>Community Tapestry</em> dashboard was presented at <strong>SANER 2026</strong> as part of the Journal First track. This tool helps open source project leaders track diversity and turnover in their communities.", "", "", "", ""),
        ("Mar 2026", "2026-03", "Dr. Guizani received funding from the <strong>Ontario Research Fund</strong> as part of Ontario's $47M investment in research and innovation, supporting our work on sustainable open source and developer productivity.", "", "", "", ""),
        ("Feb 2026", "2026-02", "Our systematic review on the impact of LLM assistants on developer productivity by Amr Mohamed, Maram Assi, and Mariam Guizani has been accepted at <strong>TOSEM</strong>.", "https://arxiv.org/abs/2507.03156", "Read the preprint", "", ""),
        ("Nov 2025", "2025-11", "Mohamad Ashraf presented our paper on reverse engineering user stories from code using LLMs at <strong>CASCON 2025</strong> in Toronto.", "https://doi.org/10.1109/CASCON66301.2025.00080", "Read the paper", "", ""),
        ("Nov 2025", "2025-11", "Michael Zhang presented our work on programming language topic classification at <strong>CASCON 2025</strong> in Toronto.", "", "", "", ""),
        ("Sep 2025", "2025-09", "The MUSE Lab is welcoming international undergraduate students through the <strong>Mitacs Globalink Research Internship</strong> program for summer 2026.", "", "", "", ""),
        ("Jul 2025", "2025-07", "Dr. Guizani was voted the favorite 3rd-year professor in Electrical and Computer Engineering at Queen's and received the <strong>Undergraduate Teaching Award</strong>.", "", "", "", ""),
        ("Jun 2025", "2025-06", "Dr. Guizani received an <strong>NSERC Discovery Grant</strong> to develop actionable diversity and inclusion tools for socio-technically sustainable open source software.", "", "", "", ""),
        ("May 2025", "2025-05", "Dr. Guizani served as Media Co-Chair for <strong>CASCON 2025</strong>, the 35th anniversary of the IEEE International Conference on Collaborative Advances in Software and Computing.", "", "", "", ""),
        ("Apr 2024", "2024-04", "Dr. Guizani received the <strong>Frank Knox Award</strong> for Excellence in Teaching at Queen's University, recognizing her outstanding commitment to students.", "", "", "", ""),
        ("Nov 2022", "2022-11", "Our research on the Maintainer Dashboard helped inform the <strong>GitHub Discussions Dashboard</strong> and was featured on the GitHub blog. This work came out of Dr. Guizani's internship at Microsoft Research.", "", "", "", ""),
        ("Sep 2021", "2021-09", "Dr. Guizani was an invited speaker at the <strong>Open Source Summit 2021</strong> in Seattle, organized by the Linux Foundation, where she presented on developer productivity in open source communities.", "", "", "", ""),
    ]
    for i, row in enumerate(existing, start=2):
        for j, value in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if j == 3:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 36

    wb.save(NEWS_XLSX)
    print(f"  created {NEWS_XLSX.relative_to(ROOT)} (pre-filled with current news)")


def init_member_template_dir():
    template_xlsx = MEMBER_TEMPLATE_DIR / "info.xlsx"
    if template_xlsx.exists():
        return
    MEMBER_TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Member"

    headers = ["Field", "Value", "Notes"]
    style_header_row(ws, headers, set())
    autosize(ws, [22, 55, 60])

    fields = [
        ("Name", "", "Full display name (required)."),
        ("Status", "Graduate", "Pick one: " + ", ".join(STATUS_CODES)),
        ("Role", "", "Job title shown under the name. Example: MASc (Fall 2024 - Pres)"),
        ("Email", "", "Contact email, or TBD."),
        ("Hobbies", "", "Comma-separated."),
        ("Photo Filename", "", "Drop your photo in this same folder and enter just the filename here (e.g. me.jpg)."),
        ("Joined", "", "Optional. YYYY-MM. Used for sorting alumni only."),
        ("Left", "", "Optional. YYYY-MM. If set, the member is treated as alumni."),
    ]
    for i, (field, value, note) in enumerate(fields, start=2):
        ws.cell(row=i, column=1, value=field).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)
        ws.cell(row=i, column=3, value=note).font = INSTR_FONT
        if field == "Status":
            dv = DataValidation(
                type="list", formula1='"' + ",".join(STATUS_CODES) + '"', allow_blank=False
            )
            dv.error = "Pick one of: " + ", ".join(STATUS_CODES)
            ws.add_data_validation(dv)
            dv.add(f"B{i}")

    # README sheet
    readme = wb.create_sheet("README")
    lines = [
        "How to add a new member:",
        "",
        "1. Copy the entire updates/members/_template/ folder.",
        "2. Rename the copy to the person's name (lowercase, hyphens), e.g. updates/members/alice-smith/.",
        "3. Open info.xlsx, fill in the Value column.",
        "4. Drop their photo file into the same folder.",
        "5. Run: python scripts/update_site.py",
        "",
        "To remove or edit a member: delete or edit their folder, then run the script.",
    ]
    for i, line in enumerate(lines, start=1):
        readme.cell(row=i, column=1, value=line).font = INSTR_FONT
    readme.column_dimensions["A"].width = 90

    wb.save(template_xlsx)
    print(f"  created {template_xlsx.relative_to(ROOT)}")


def export_existing_members():
    """First run only: export current members.json into per-folder Excel files
    so the user has a starting point to edit."""
    if not MEMBERS_JSON.exists():
        return
    members = load_json(MEMBERS_JSON, [])

    # Map page data (in members.html) into structured entries
    page_data = {
        "Mariam Guizani":   {"status": "PI",            "role": "Assistant Professor & Connected Minds Member", "email": "mariam.guizani@queensu.ca", "hobbies": "Running, Hiking", "photo": "mariamguizani.jpeg"},
        "Robyn Elliot":     {"status": "Graduate",      "role": "PhD (Fall 2024 - pres)",     "email": "5rje@queensu.ca",  "hobbies": "Playing the bagpipes, martial arts", "photo": "eliottrobyn.jpeg"},
        "Mohamed Ouf":      {"status": "Graduate",      "role": "MASc (Fall 2024 - Pres)",    "email": "24blr2@queensu.ca","hobbies": "Swimming, Strategies/Card Games",     "photo": "mohamedouf.jpeg"},
        "Amr Mohamed":      {"status": "Graduate",      "role": "MASc (Fall 2024 - Pres)",    "email": "amr.m@queensu.ca", "hobbies": "Chess, Piano, Guitar",                "photo": "amrmohamed.jpg"},
        "Michael Zhang":    {"status": "Graduate",      "role": "MEng (Mar 2024 - Pres)",     "email": "18jz111@queensu.ca","hobbies": "Skiing, camping, hiking, card and board games", "photo": "michaelzhang.jpeg"},
        "Aaliyah Chang":    {"status": "Graduate",      "role": "MASc Student in Electrical and Computer Engineering", "email": "TBD", "hobbies": "TBD", "photo": "aaliyah-chang.png"},
        "Yawar Ashraf":     {"status": "Graduate",      "role": "MASc Student",               "email": "TBD",              "hobbies": "TBD",                                  "photo": "yawar-ashraf.png"},
        "Zeph Van Iterson": {"status": "Undergraduate", "role": "Research Assistant (Mar 2024 - pres)", "email": "20zrvi@queensu.ca", "hobbies": "Cooking, Movies",          "photo": "zephvaniterson.jpg"},
        "Haoyu Li":         {"status": "Alumni",        "role": "MEng (Mar 2024 - Nov 2024)", "email": "17hl66@queensu.ca","hobbies": "",                                    "photo": "default.jpg"},
        "Elias Frigui":     {"status": "Alumni",        "role": "Research Assistant (Mar 2024 - Aug 2024)", "email": "21ef32@queensu.ca", "hobbies": "",                     "photo": "default.jpg"},
    }

    for name, info in page_data.items():
        folder_name = slugify(name)
        folder = MEMBERS_DIR / folder_name
        if folder.exists():
            continue
        folder.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Member"
        style_header_row(ws, ["Field", "Value", "Notes"], set())
        autosize(ws, [22, 55, 60])
        rows = [
            ("Name", name, ""),
            ("Status", info["status"], "Pick one: " + ", ".join(STATUS_CODES)),
            ("Role", info["role"], ""),
            ("Email", info["email"], ""),
            ("Hobbies", info["hobbies"], ""),
            ("Photo Filename", info["photo"], "Copy the photo into this folder if you change it."),
            ("Joined", "", ""),
            ("Left", "", ""),
        ]
        for i, (field, value, note) in enumerate(rows, start=2):
            ws.cell(row=i, column=1, value=field).font = Font(bold=True)
            ws.cell(row=i, column=2, value=value)
            ws.cell(row=i, column=3, value=note).font = INSTR_FONT
            if field == "Status":
                dv = DataValidation(
                    type="list", formula1='"' + ",".join(STATUS_CODES) + '"', allow_blank=False
                )
                ws.add_data_validation(dv)
                dv.add(f"B{i}")
        wb.save(folder / "info.xlsx")
        # Copy photo into the folder for reference
        photo_src = PHOTOS / info["photo"]
        photo_dst = folder / info["photo"]
        if photo_src.exists() and not photo_dst.exists():
            shutil.copy(photo_src, photo_dst)
        print(f"  exported {folder.relative_to(ROOT)}/")


def init_all():
    UPDATES.mkdir(parents=True, exist_ok=True)
    init_publications_template()
    init_news_template()
    init_member_template_dir()
    export_existing_members()


# --------------------------------------------------------------------------
# Process: Publications
# --------------------------------------------------------------------------

def process_publications():
    if not PUBS_XLSX.exists():
        return False
    wb = load_workbook(PUBS_XLSX, data_only=True)
    if "Publications" not in wb.sheetnames:
        return False
    ws = wb["Publications"]

    headers = [(c.value or "").rstrip(" *") for c in ws[1]]
    new_entries = []
    today = date.today().isoformat()

    for row_idx in range(2, ws.max_row + 1):
        row = {h: ws.cell(row=row_idx, column=i + 1).value for i, h in enumerate(headers)}
        title = (row.get("Title") or "").strip()
        if not title:
            continue
        authors = split_list(row.get("Authors"))
        year = row.get("Year")
        venue = (row.get("Venue") or "").strip()
        ptype = (row.get("Type") or "").strip().lower()
        topics = [t.lower() for t in split_list(row.get("Topics"))]
        unknown_topics = [t for t in topics if t not in TOPIC_CODES]
        if unknown_topics:
            print(f"  row {row_idx}: unknown topics {unknown_topics} — skipping")
            continue
        if ptype not in TYPE_CODES:
            print(f"  row {row_idx}: invalid type {ptype!r} — skipping")
            continue
        if not (authors and year and venue):
            print(f"  row {row_idx}: missing required field — skipping")
            continue

        doi = (row.get("DOI") or "").strip().replace("https://doi.org/", "")
        venue_acronym = (row.get("Venue Acronym") or "").strip()
        pdf = (row.get("PDF URL") or "").strip()
        slides = (row.get("Slides URL") or "").strip()
        abstract = (row.get("Abstract") or "").strip()

        # Slides may be a local filename; copy it to files/
        if slides and "://" not in slides:
            src = PUBS_FILES_DIR / slides
            if src.exists():
                FILES.mkdir(parents=True, exist_ok=True)
                shutil.copy(src, FILES / slides)
                slides = "files/" + slides
            else:
                print(f"  row {row_idx}: slides file {slides!r} not found in updates/publications/files/")
                slides = ""

        links = {}
        if doi:
            links["doi"] = "https://doi.org/" + doi
        if pdf:
            links["pdf"] = pdf
        if slides:
            links["presentation"] = slides

        entry = {
            "id": doi or slugify(f"{authors[0]} {year} {title.split()[0]}"),
            "title": title,
            "authors": authors,
            "venue": venue,
            "venue_acronym": venue_acronym,
            "year": int(year),
            "type": ptype,
            "topics": topics,
            "doi": doi,
            "links": links,
            "added_by": "manual",
            "added_on": today,
            "abstract": abstract,
        }
        new_entries.append(entry)

    existing = load_json(PUBS_JSON, [])
    by_key = {(p.get("doi") or p["title"].lower().strip()): p for p in existing}

    added = 0
    updated = 0
    for entry in new_entries:
        key = entry["doi"] or entry["title"].lower().strip()
        if key in by_key:
            by_key[key].update(entry)
            updated += 1
        else:
            existing.append(entry)
            by_key[key] = entry
            added += 1

    write_json(PUBS_JSON, existing)
    print(f"  publications: {added} added, {updated} updated, total {len(existing)}")
    return True


# --------------------------------------------------------------------------
# Process: News
# --------------------------------------------------------------------------

def process_news():
    if not NEWS_XLSX.exists():
        return False
    wb = load_workbook(NEWS_XLSX, data_only=True)
    ws = wb["News"] if "News" in wb.sheetnames else wb.active

    items = []
    for row_idx in range(2, ws.max_row + 1):
        badge = ws.cell(row=row_idx, column=1).value
        sort_date = ws.cell(row=row_idx, column=2).value
        text = ws.cell(row=row_idx, column=3).value
        url = ws.cell(row=row_idx, column=4).value
        label = ws.cell(row=row_idx, column=5).value
        image = ws.cell(row=row_idx, column=6).value
        alt = ws.cell(row=row_idx, column=7).value
        if not (badge and text):
            continue
        if image:
            image = str(image).strip()
            src = NEWS_IMAGES_DIR / image
            if src.exists():
                PHOTOS.mkdir(parents=True, exist_ok=True)
                shutil.copy(src, PHOTOS / image)
        items.append({
            "badge": str(badge).strip(),
            "sort_date": str(sort_date).strip() if sort_date else "",
            "text": str(text).strip(),
            "link_url": str(url).strip() if url else "",
            "link_label": str(label).strip() if label else "",
            "image": str(image).strip() if image else "",
            "image_alt": str(alt).strip() if alt else "",
        })

    items.sort(key=lambda x: x["sort_date"], reverse=True)
    write_json(NEWS_JSON, items)

    html_blocks = []
    for it in items:
        block = ['                    <div class="news-item">']
        block.append(f'                        <span class="news-badge">{it["badge"]}</span>')
        block.append(f'                        <p>{it["text"]}</p>')
        if it["link_url"]:
            label = it["link_label"] or "Learn more"
            block.append(
                f'                        <a href="{it["link_url"]}" class="news-link" target="_blank">{label}</a>'
            )
        if it["image"]:
            alt = it["image_alt"] or ""
            block.append(
                f'                        <img src="photos/{it["image"]}" class="news-photo" alt="{alt}">'
            )
        block.append('                    </div>')
        html_blocks.append("\n".join(block))

    rendered = "\n".join(html_blocks)
    text = INDEX_HTML.read_text()
    text = replace_between_markers(
        text,
        "<!-- AUTO:NEWS:START -->",
        "<!-- AUTO:NEWS:END -->",
        rendered,
    )
    INDEX_HTML.write_text(text)
    print(f"  news: {len(items)} items written to index.html and data/news.json")
    return True


# --------------------------------------------------------------------------
# Process: Members
# --------------------------------------------------------------------------

def read_member_folder(folder):
    info_path = folder / "info.xlsx"
    if not info_path.exists():
        return None
    wb = load_workbook(info_path, data_only=True)
    ws = wb["Member"] if "Member" in wb.sheetnames else wb.active
    data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        data[str(row[0]).strip()] = (row[1] if len(row) > 1 else "") or ""
    name = (data.get("Name") or "").strip()
    if not name:
        return None
    status = (data.get("Status") or "Graduate").strip()
    if status not in STATUS_CODES:
        print(f"  {folder.name}: invalid status {status!r}, defaulting to Graduate")
        status = "Graduate"
    photo = (data.get("Photo Filename") or "").strip()
    if photo:
        src = folder / photo
        if src.exists():
            PHOTOS.mkdir(parents=True, exist_ok=True)
            shutil.copy(src, PHOTOS / photo)
        elif not (PHOTOS / photo).exists():
            print(f"  {folder.name}: photo {photo!r} not found, using default")
            photo = "default.jpg"
    else:
        photo = "default.jpg"
    return {
        "name": name,
        "status": status,
        "role": (data.get("Role") or "").strip(),
        "email": (data.get("Email") or "").strip(),
        "hobbies": (data.get("Hobbies") or "").strip(),
        "photo": photo,
        "joined": (data.get("Joined") or "").strip(),
        "left": (data.get("Left") or "").strip(),
    }


def render_member_card(m, is_pi=False):
    card_class = "card principal-investigator-card h-100" if is_pi else "card h-100"
    lines = [
        '           <div class="col">' if not is_pi else '       <div class="col-md-4">',
        f'               <div class="{card_class}">',
        f'                   <img src="photos/{m["photo"]}" class="card-img-top" alt="Photo of {m["name"]}">',
        '                   <div class="card-body">',
        f'                       <h5 class="card-title">{m["name"]}</h5>',
        f'                       <h6 class="card-subtitle mb-2 text-muted">{m["role"]}</h6>',
    ]
    if m["email"]:
        lines.append(f'                       <p class="card-text">Email: {m["email"]}</p>')
    if m["hobbies"]:
        lines.append(f'                       <p class="card-text">Hobbies: <i>{m["hobbies"]}</i></p>')
    lines.append('                   </div>')
    lines.append('               </div>')
    lines.append('           </div>')
    return "\n".join(lines)


def render_member_section(title, members):
    count = len(members)
    label = "member" if count == 1 else "members"
    out = [
        '<div class="container">',
        '   <div class="role-heading mt-5 mb-4">',
        f'   <h1 class="page-title mb-0">{title}</h1>',
        f'   <span class="role-count">{count} {label}</span>',
        '   </div>',
        '   <div class="row row-cols-1 row-cols-md-5 g-4">',
    ]
    for m in members:
        out.append(render_member_card(m))
    out.append('   </div>')
    out.append('</div>')
    return "\n".join(out)


def process_members():
    if not MEMBERS_DIR.exists():
        return False
    members = []
    for folder in sorted(MEMBERS_DIR.iterdir()):
        if not folder.is_dir() or folder.name.startswith("_"):
            continue
        m = read_member_folder(folder)
        if m:
            members.append(m)

    if not members:
        print("  members: no member folders found")
        return False

    # Preserve aliases and semantic_scholar_ids from existing JSON
    existing_meta = {m.get("name"): m for m in load_json(MEMBERS_JSON, [])}
    write_json(MEMBERS_JSON, [
        {
            "name": m["name"],
            "aliases": existing_meta.get(m["name"], {}).get("aliases", []),
            "role": m["role"],
            "status": m["status"],
            "semantic_scholar_ids": existing_meta.get(m["name"], {}).get("semantic_scholar_ids", []),
            "active": m["status"] != "Alumni",
        }
        for m in members
    ])

    pi = [m for m in members if m["status"] == "PI"]
    grad = [m for m in members if m["status"] == "Graduate"]
    undergrad = [m for m in members if m["status"] == "Undergraduate"]
    alumni = [m for m in members if m["status"] == "Alumni"]

    # Build the full block
    blocks = []
    if pi:
        m = pi[0]
        blocks.append(
            '<div class="container">\n'
            '  <div class="role-heading mt-4 mb-4">\n'
            '  <h1 class="page-title mb-0">Principal Investigator</h1>\n'
            '  <span class="role-count">1 member</span>\n'
            '  </div>\n'
            '  <p class="page-intro">Meet the researchers, students, and contributors who are shaping work at MUSE Lab.</p>\n'
            '  <div class="row g-4">\n'
            + render_member_card(m, is_pi=True) + '\n'
            '  </div>\n'
            '</div>'
        )
    if grad:
        blocks.append(render_member_section("Graduate Students", grad))
    if undergrad:
        blocks.append(render_member_section("Undergraduate Students", undergrad))
    if alumni:
        blocks.append(render_member_section("Alumni", alumni))

    rendered = "\n\n".join(blocks)
    text = MEMBERS_HTML.read_text()
    text = replace_between_markers(
        text,
        "<!-- AUTO:MEMBERS:START -->",
        "<!-- AUTO:MEMBERS:END -->",
        rendered,
    )
    MEMBERS_HTML.write_text(text)
    print(f"  members: {len(members)} cards written to members.html and data/members.json")
    return True


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--init-only",
        action="store_true",
        help="Only create the templates in updates/, do not process anything.",
    )
    args = parser.parse_args()

    first_run = not UPDATES.exists() or not any(UPDATES.iterdir())
    if first_run or args.init_only:
        print("Setting up updates/ folder...")
        init_all()
        if args.init_only:
            print("Done. Edit the files in updates/ and run the script again without --init-only.")
            return
        print("")

    print("Processing updates...")
    init_all()  # make sure templates exist
    process_publications()
    process_news()
    process_members()
    print("\nAll done. Run `git status` to see what changed, then commit.")


if __name__ == "__main__":
    main()
